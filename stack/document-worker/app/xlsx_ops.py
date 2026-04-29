from io import BytesIO
import random
from openpyxl import load_workbook
from openpyxl.utils.cell import range_boundaries


def _resolve_worksheet(wb, update: dict):
    sheet = update.get("sheet")
    if isinstance(sheet, str) and sheet.strip():
        return wb[sheet.strip()]
    if not wb.sheetnames:
        raise ValueError("Workbook has no sheets.")
    return wb[wb.sheetnames[0]]


def _fill_range_with_value(ws, cell_range: str, value):
    min_col, min_row, max_col, max_row = range_boundaries(cell_range)
    for row in range(min_row, max_row + 1):
        for col in range(min_col, max_col + 1):
            ws.cell(row=row, column=col).value = value


def _fill_range_with_random_money(ws, cell_range: str, update: dict):
    min_col, min_row, max_col, max_row = range_boundaries(cell_range)

    min_value = float(update.get("min", 1000))
    max_value = float(update.get("max", 100000))
    if min_value > max_value:
        min_value, max_value = max_value, min_value

    decimals = int(update.get("decimals", 2))
    if decimals < 0:
        decimals = 0
    if decimals > 6:
        decimals = 6

    for row in range(min_row, max_row + 1):
        for col in range(min_col, max_col + 1):
            ws.cell(row=row, column=col).value = round(
                random.uniform(min_value, max_value), decimals
            )


def xlsx_update_cells(xlsx_bytes: bytes, updates: list[dict]) -> bytes:
    """
    Supported update formats:
    - {"sheet":"Sheet1","cell":"B2","value":"123"}
    - {"sheet":"Sheet1","range":"A2:A500","value":"foo"}
    - {"sheet":"Sheet1","range":"A2:A500","generator":"random_money","min":1000,"max":100000,"decimals":2}
    Sheet is optional; if missing, first worksheet is used.
    """
    if not isinstance(updates, list) or len(updates) == 0:
        raise ValueError("updates must be a non-empty list.")

    wb = load_workbook(filename=BytesIO(xlsx_bytes))

    for idx, u in enumerate(updates):
        if not isinstance(u, dict):
            raise ValueError(f"updates[{idx}] must be an object.")

        ws = _resolve_worksheet(wb, u)

        if "cell" in u:
            cell_ref = str(u.get("cell", "")).strip()
            if not cell_ref:
                raise ValueError(f"updates[{idx}].cell must be non-empty.")
            ws[cell_ref].value = u.get("value")
            continue

        if "range" in u:
            cell_range = str(u.get("range", "")).strip()
            if not cell_range:
                raise ValueError(f"updates[{idx}].range must be non-empty.")

            generator = str(u.get("generator", "")).strip().lower()
            if generator in ("random_money", "random_amount", "random_currency"):
                _fill_range_with_random_money(ws, cell_range, u)
            else:
                _fill_range_with_value(ws, cell_range, u.get("value"))
            continue

        raise ValueError(
            f"updates[{idx}] requires either 'cell' or 'range'."
        )

    out = BytesIO()
    wb.save(out)
    return out.getvalue()
